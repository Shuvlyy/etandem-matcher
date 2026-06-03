import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


def export_results_to_excel(df: pd.DataFrame, filename: str = "resultats.xlsx"):
    """Exports results to a new Excel file."""
    print(f"\Exporting results to {filename}...")

    df.to_excel(filename, index=False, engine="openpyxl")

    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.title = "Match Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2C3E50", end_color="2C3E50", fill_type="solid"
    )
    center_aligned_text = Alignment(horizontal="center", vertical="center")

    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text

        column_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 30

    for row_idx in range(2, len(df) + 2):
        ws.cell(row=row_idx, column=3).alignment = center_aligned_text
        ws.cell(row=row_idx, column=4).alignment = center_aligned_text

    wb.save(filename)
